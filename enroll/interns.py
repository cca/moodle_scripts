import _csv  # for typing
import csv
import re
import warnings
from pathlib import Path
from typing import Any, Generator, Literal

import click
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

program_to_course_map: dict[str, str] = {
    "Architecture": "BARCH-INTRN",
    "Graduate Architecture": "MARCH-INTRN",
    "Graphic Design": "GRAPH-INTRN",
    "Industrial Design": "INDUS-INTRN",
    "Interaction Design": "IXDSN-INTRN",
    "Interior Design": "INTER-INTRN",
}
programs_with_internship: list[str] = list(program_to_course_map.keys())


def row_to_dict(header, row) -> dict[Any, Any]:
    return dict(zip(header, row))


def meets_program_criteria(student: dict[str, Any], program: str | None = None) -> bool:
    """Check if student meets the criteria to be added to their major's
    internship course. Criteria rely on the Latest Class Standing field, which
    is a string that goes from "First Year" up to "Fifth Year". It is not a
    _level_ per se, as in an undergrad can be a "Fifth Year" and grad students
    start in their "First Year". "First Year" != Freshman.

    Args:
        student (dict): dict of student information

    Returns:
        bool: true if ready, false otherwise
    """
    major: str = student["Primary Program of Study"]
    level: str = student["Latest Class Standing"]
    # INDUS wants students to finish Prof Practice, we do not preload them
    # IXDSN has their own "student tracking" spreadsheet we use
    # but if we're using program filter, always return students in that program
    # even if it is not one of the ones we normally preload
    programs = [
        "Architecture",
        "Interior Design",
        "Graphic Design",
    ]
    if program:
        programs.append(program)

    if major in programs and level == "Third Year":
        return True
    if major == "Industrial Design" and level == "Fourth Year":
        return True
    if major == "Graduate Architecture" and level == "Second Year":
        return True
    return False


def make_enrollments(student, semester, program=None, list_mode=False) -> list[Any]:
    """return enrollment rows if student meets general criteria
    if program is present, only returns rows for students in that program

    Args:
        student (dict): dict of student information
        program (str|None): program filter or None for all programs

    Returns:
        list: returns a list ready to be added to a Moodle enrollment CSV
        if the student is ready, otherwise the boolean False
    """
    # students must be actively enrolled in a program with a required internship
    major: str = student["Primary Program of Study"]
    if (
        major in programs_with_internship
        and student["Primary Program of Study Record Status"] == "In Progress"
        and student["CCA Email"]
    ):
        if program and major != program:
            return []
        if not meets_program_criteria(student, program):
            return []
        # return enrollment rows
        username: str = re.sub("@cca.edu", "", student["CCA Email"])
        course: str = program_to_course_map[major]
        is_intl: Literal["International", False] = (
            "International" if student["Is International Student"] == "Yes" else False
        )
        if list_mode:
            return [student["Student"], student["CCA Email"]]
        if is_intl:
            # intl students need 2 enrollments so they can be in 2 groups (semester and intl)
            return [(username, course, semester), (username, course, is_intl)]
        return [(username, course, semester)]
    return []


def wd_report_to_enroll_csv(
    report: Path, semester: str, program: str, list_mode: bool
) -> None:
    # silence "Workbook contains no default style" warning
    with warnings.catch_warnings(record=True):
        warnings.simplefilter("always")
        wb: Workbook = load_workbook(report)
    sheet: Worksheet = wb.worksheets[0]
    rows: Generator = sheet.iter_rows(values_only=True)
    header: tuple = next(rows)
    if header[0] == "Students for Internship Review":
        # skip header row
        header = next(rows)
    with open("enrollments.csv", "w") as file:
        writer: _csv._writer = csv.writer(file)
        # write CSV header row
        writer.writerow(["username", "course1", "group1"])
        if list_mode:
            click.echo("\t".join(["Student", "Email"]))
        for row in rows:
            student: dict[str, str] = row_to_dict(header, row)
            enrollments: list[Any] = make_enrollments(
                student, semester, program, list_mode
            )
            if list_mode and len(enrollments):
                click.echo("\t".join(enrollments))
            else:
                writer.writerows(enrollments)


def semester_validator(ctx, param, value):
    """validate semester string for click"""
    if re.match(r"(Spring|Fall|Summer) \d{4}", value):
        return value
    raise click.BadParameter(
        f"Semester must be in the format of 'Season YYYY' like 'Fall 2023', not '{value}'"
    )


@click.command(
    help="Generate enrollments for students who are ready for internship courses."
)
@click.help_option("-h", "--help")
@click.option(
    "-r",
    "--report",
    default="data/Students_for_Internship_Review.xlsx",
    help="path to the Workday Excel file (relative to project root)",
    required=True,
    type=click.Path(exists=True, path_type=Path),
)
@click.option(
    "-s",
    "--semester",
    callback=semester_validator,
    help='semester group (like "Fall 2023"))',
    required=True,
)
@click.option(
    "-p",
    "--program",
    help="Generate enrollments for only a specific program",
    type=click.Choice(programs_with_internship),
)
@click.option(
    "-l",
    "--list-mode",
    is_flag=True,
    help="print list of students (instead of CSV)",
)
def main(report: Path, semester: str, program: str, list_mode: bool):
    wd_report_to_enroll_csv(report, semester, program, list_mode)
    if not list_mode:
        click.echo(
            "Created enrollments.csv. Upload Users: https://moodle.cca.edu/admin/tool/uploaduser/"
        )
        click.echo(
            f"Remember to add {semester} to the Semester Groups grouping in each course."
        )


if __name__ == "__main__":
    main()
